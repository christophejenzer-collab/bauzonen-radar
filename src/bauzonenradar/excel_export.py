"""
excel_export.py - XLSX-Export der Iter-5-Massen-Analyse.

Liest aus KantonsCache und erstellt eine Excel-Datei mit 6 Sheets:
    Statistik         - Counts pro Kategorie und Zone, Top-Zahlen
    Verdichtung       - sortiert nach Reserve absteigend
    Neugeschaeft      - sortiert nach Theoretisch zulaessig
    Ersatzneubau      - sortiert nach Reserve
    Ausgereizt        - alle ausgereizten (Bestandsschutz-Faelle)
    Alle Parzellen    - komplette Tabelle mit Filter

CLI:
    python excel_export.py "Oberhofen am Thunersee"
    python excel_export.py "Oberhofen am Thunersee" --kanton be --ausgabe ausgaben/

Iter 5 | Stand: 12. Mai 2026
"""

from __future__ import annotations

import argparse
import datetime
import logging
import sys
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------

DEFAULT_KANTON = "be"
DEFAULT_AUSGABE_DIR = Path("ausgaben")
CACHE_DIR = Path("cache")

GRUDIS_BASIS_URL = (
    "https://www.geo.apps.be.ch/de/applikationen/"
    "grundstuecksinformation.html?egrid={egrid}"
)

# Header-Styling
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")


# ---------------------------------------------------------------------------
# DB-Zugriff
# ---------------------------------------------------------------------------

def lade_parzellen_aus_cache(
    kanton: str,
    gemeinde: str,
    cache_dir: Path = CACHE_DIR,
) -> list[dict]:
    """Laedt alle Parzellen einer Gemeinde aus dem SQLite-Cache.

    Returns:
        Liste von dicts mit allen flachen Spalten der parzellen-Tabelle.
    """
    db_pfad = cache_dir / f"cache_{kanton.lower()}.db"
    if not db_pfad.exists():
        raise FileNotFoundError(f"Cache-Datei nicht gefunden: {db_pfad}")

    conn = sqlite3.connect(str(db_pfad))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT
            egrid, gemeinde, bfs_nummer,
            parzellen_nummer, parzellen_flaeche_m2,
            datenqualitaet, zone, theoretisch_zulaessig_m2,
            gwr_summe_geschossflaeche_m2, reserve_m2, reserve_prozent,
            ausschoepfungsgrad_prozent, anzahl_gebaeude,
            klassifikation, fehler, abfrage_datum
        FROM parzellen
        WHERE gemeinde = ?
        ORDER BY parzellen_nummer
    """, (gemeinde,))
    parzellen = [dict(row) for row in cur.fetchall()]
    conn.close()
    return parzellen


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _grudis_link(egrid: str) -> str:
    """Liefert GRUDIS-URL fuer einen EGRID."""
    return GRUDIS_BASIS_URL.format(egrid=egrid)


def _try_int(wert) -> int | str:
    """Konvertiert zu int falls moeglich, sonst leer."""
    if wert is None:
        return ""
    try:
        return int(wert)
    except (ValueError, TypeError):
        return ""


def _try_float(wert) -> float | str:
    """Konvertiert zu float falls moeglich, sonst leer."""
    if wert is None:
        return ""
    try:
        return round(float(wert), 1)
    except (ValueError, TypeError):
        return ""


def _formatiere_header(sheet, header_zeile: int = 1) -> None:
    """Stylet die Header-Zeile."""
    for cell in sheet[header_zeile]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _autosize_spalten(sheet, max_breite: int = 30) -> None:
    """Setzt Spaltenbreite ungefaehr passend zum Inhalt."""
    for spalten_index, spalte in enumerate(sheet.columns, start=1):
        max_laenge = 0
        for cell in spalte:
            try:
                wert = str(cell.value) if cell.value is not None else ""
                if len(wert) > max_laenge:
                    max_laenge = len(wert)
            except Exception:
                pass
        breite = min(max_laenge + 2, max_breite)
        sheet.column_dimensions[get_column_letter(spalten_index)].width = max(8, breite)


def _freeze_header(sheet) -> None:
    """Friert die erste Datenzeile ein."""
    sheet.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

PARZELLEN_HEADER = [
    "Parz. Nr.", "EGRID", "Zone", "Flaeche (m²)",
    "Soll (m²)", "Ist GWR (m²)", "Reserve (m²)", "Reserve %",
    "Geb.", "Datenqualitaet", "Klassifikation", "GRUDIS",
]


def _parzelle_zu_row(p: dict) -> list:
    """Konvertiert ein Parzellen-dict in eine Excel-Row."""
    return [
        p.get("parzellen_nummer") or "",
        p.get("egrid") or "",
        p.get("zone") or "",
        _try_int(p.get("parzellen_flaeche_m2")),
        _try_int(p.get("theoretisch_zulaessig_m2")),
        _try_int(p.get("gwr_summe_geschossflaeche_m2")),
        _try_float(p.get("reserve_m2")),
        _try_float(p.get("reserve_prozent")),
        _try_int(p.get("anzahl_gebaeude")),
        p.get("datenqualitaet") or "",
        p.get("klassifikation") or "",
        _grudis_link(p["egrid"]) if p.get("egrid") else "",
    ]


def _schreibe_parzellen_sheet(
    sheet,
    parzellen: list[dict],
    titel: str,
    untertitel: str = "",
) -> None:
    """Schreibt eine Liste von Parzellen als formatierte Tabelle."""
    # Titel-Zeile
    sheet["A1"] = titel
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PARZELLEN_HEADER))

    if untertitel:
        sheet["A2"] = untertitel
        sheet["A2"].font = SUBTITLE_FONT
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(PARZELLEN_HEADER))
        start_zeile = 4
    else:
        start_zeile = 3

    # Header
    for spalten_index, header in enumerate(PARZELLEN_HEADER, start=1):
        cell = sheet.cell(row=start_zeile, column=spalten_index, value=header)
    _formatiere_header(sheet, start_zeile)

    # Daten
    for zeilen_index, p in enumerate(parzellen, start=start_zeile + 1):
        row = _parzelle_zu_row(p)
        for spalten_index, wert in enumerate(row, start=1):
            cell = sheet.cell(row=zeilen_index, column=spalten_index, value=wert)
            cell.border = THIN_BORDER
            # GRUDIS-Link als Hyperlink
            if spalten_index == len(PARZELLEN_HEADER) and wert:
                cell.hyperlink = wert
                cell.value = "→ GRUDIS"
                cell.font = Font(color="0563C1", underline="single")

    _autosize_spalten(sheet)
    sheet.freeze_panes = sheet.cell(row=start_zeile + 1, column=1).coordinate


def _sheet_statistik(wb: Workbook, parzellen: list[dict], gemeinde: str, kanton: str) -> None:
    """Erstellt Sheet 'Statistik' mit Counts und Uebersicht."""
    sheet = wb.create_sheet("Statistik", 0)  # erstes Sheet

    sheet["A1"] = f"Bauzonen-Radar: Massen-Analyse {gemeinde}"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:D1")

    sheet["A2"] = (
        f"Kanton: {kanton.upper()} | "
        f"Stand: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')} | "
        f"Total Parzellen: {len(parzellen)}"
    )
    sheet["A2"].font = SUBTITLE_FONT
    sheet.merge_cells("A2:D2")

    # Klassifikation
    sheet["A4"] = "Klassifikation"
    sheet["A4"].font = Font(bold=True, size=12)

    counts: dict[str, int] = {}
    for p in parzellen:
        k = p.get("klassifikation") or "?"
        counts[k] = counts.get(k, 0) + 1

    reihenfolge = [
        ("VERDICHTUNG", "Verdichtung (bebaut + Reserve)"),
        ("NEUGESCHAEFT", "Neugeschaeft (leer + Bauland)"),
        ("ERSATZNEUBAU", "Ersatzneubau (alt + Reserve)"),
        ("UNAUFFAELLIG", "Unauffaellig (wenig Reserve)"),
        ("AUSGEREIZT", "Ausgereizt (Bestandsschutz)"),
        ("AUSSCHLUSS_REGLEMENT", "Ausschluss: Reglement-Sperre"),
        ("AUSSCHLUSS_ZU_KLEIN", "Ausschluss: Parzelle zu klein"),
        ("AUSSCHLUSS_FEHLER", "Ausschluss: Analyse-Fehler"),
    ]

    sheet.cell(row=5, column=1, value="Kategorie").font = HEADER_FONT
    sheet.cell(row=5, column=1).fill = HEADER_FILL
    sheet.cell(row=5, column=2, value="Beschreibung").font = HEADER_FONT
    sheet.cell(row=5, column=2).fill = HEADER_FILL
    sheet.cell(row=5, column=3, value="Anzahl").font = HEADER_FONT
    sheet.cell(row=5, column=3).fill = HEADER_FILL
    sheet.cell(row=5, column=4, value="Anteil").font = HEADER_FONT
    sheet.cell(row=5, column=4).fill = HEADER_FILL

    total = len(parzellen) or 1
    zeile = 6
    for key, label in reihenfolge:
        anzahl = counts.get(key, 0)
        anteil = anzahl / total * 100
        sheet.cell(row=zeile, column=1, value=key)
        sheet.cell(row=zeile, column=2, value=label)
        sheet.cell(row=zeile, column=3, value=anzahl)
        sheet.cell(row=zeile, column=4, value=f"{anteil:.1f}%")
        for spalte in range(1, 5):
            sheet.cell(row=zeile, column=spalte).border = THIN_BORDER
        zeile += 1

    # Top-Zonen
    zeile += 2
    sheet.cell(row=zeile, column=1, value="Top Zonen").font = Font(bold=True, size=12)
    zeile += 1
    sheet.cell(row=zeile, column=1, value="Zone").font = HEADER_FONT
    sheet.cell(row=zeile, column=1).fill = HEADER_FILL
    sheet.cell(row=zeile, column=2, value="Anzahl").font = HEADER_FONT
    sheet.cell(row=zeile, column=2).fill = HEADER_FILL
    zeile += 1

    zonen_counts: dict[str, int] = {}
    for p in parzellen:
        z = p.get("zone") or "(keine)"
        zonen_counts[z] = zonen_counts.get(z, 0) + 1

    for z, n in sorted(zonen_counts.items(), key=lambda x: -x[1])[:15]:
        sheet.cell(row=zeile, column=1, value=z)
        sheet.cell(row=zeile, column=2, value=n)
        sheet.cell(row=zeile, column=1).border = THIN_BORDER
        sheet.cell(row=zeile, column=2).border = THIN_BORDER
        zeile += 1

    # Spalten-Breite
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 40
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12


# ---------------------------------------------------------------------------
# Haupt-Funktion
# ---------------------------------------------------------------------------

def exportiere_gemeinde(
    gemeinde: str,
    kanton: str = DEFAULT_KANTON,
    ausgabe_dir: Path = DEFAULT_AUSGABE_DIR,
    cache_dir: Path = CACHE_DIR,
) -> Path:
    """Exportiert eine Gemeinde aus dem Cache nach XLSX.

    Returns:
        Pfad zur erstellten Excel-Datei.
    """
    logger.info(f"Exportiere {gemeinde} (Kanton {kanton.upper()})...")
    parzellen = lade_parzellen_aus_cache(kanton, gemeinde, cache_dir=cache_dir)
    logger.info(f"  {len(parzellen)} Parzellen aus Cache geladen")

    if not parzellen:
        raise ValueError(f"Keine Parzellen fuer Gemeinde '{gemeinde}' im Cache")

    # Filter pro Kategorie
    def filter_kat(*kategorien):
        return [p for p in parzellen if p.get("klassifikation") in kategorien]

    verdichtung = filter_kat("VERDICHTUNG")
    neugeschaeft = filter_kat("NEUGESCHAEFT")
    ersatzneubau = filter_kat("ERSATZNEUBAU")
    ausgereizt = filter_kat("AUSGEREIZT")

    # Sortierung
    def sort_reserve(p):
        return -(p.get("reserve_m2") or 0)

    def sort_zulaessig(p):
        return -(p.get("theoretisch_zulaessig_m2") or 0)

    verdichtung.sort(key=sort_reserve)
    neugeschaeft.sort(key=sort_zulaessig)
    ersatzneubau.sort(key=sort_reserve)
    ausgereizt.sort(key=sort_reserve)

    # Workbook
    wb = Workbook()
    # Default-Sheet entfernen
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: Statistik (oben)
    _sheet_statistik(wb, parzellen, gemeinde, kanton)

    # Sheet 2-5: Kategorien
    _schreibe_parzellen_sheet(
        wb.create_sheet("Verdichtung"),
        verdichtung,
        "Verdichtung - bebaute Parzellen mit Reserve",
        f"{len(verdichtung)} Parzellen, sortiert nach Reserve absteigend",
    )
    _schreibe_parzellen_sheet(
        wb.create_sheet("Neugeschaeft"),
        neugeschaeft,
        "Neugeschaeft - leere Parzellen mit Bauland",
        f"{len(neugeschaeft)} Parzellen, sortiert nach Theoretisch zulaessig absteigend",
    )
    _schreibe_parzellen_sheet(
        wb.create_sheet("Ersatzneubau"),
        ersatzneubau,
        "Ersatzneubau - alte Bauten mit Reserve",
        f"{len(ersatzneubau)} Parzellen, sortiert nach Reserve absteigend",
    )
    _schreibe_parzellen_sheet(
        wb.create_sheet("Ausgereizt"),
        ausgereizt,
        "Ausgereizt - GWR uebersteigt Soll (Bestandsschutz)",
        f"{len(ausgereizt)} Parzellen, sortiert nach Reserve",
    )

    # Sheet 6: Alle Parzellen
    _schreibe_parzellen_sheet(
        wb.create_sheet("Alle"),
        parzellen,
        f"Alle Parzellen - {gemeinde}",
        f"{len(parzellen)} Parzellen, sortiert nach Parzellen-Nummer",
    )

    # Speichern
    ausgabe_dir.mkdir(parents=True, exist_ok=True)
    gemeinde_slug = gemeinde.lower().replace(" ", "_").replace("ü", "ue").replace("ä", "ae").replace("ö", "oe")
    datum = datetime.datetime.now().strftime("%Y-%m-%d")
    dateiname = f"bauzonen_radar_{gemeinde_slug}_{datum}.xlsx"
    ausgabe_pfad = ausgabe_dir / dateiname

    wb.save(ausgabe_pfad)
    logger.info(f"  Excel gespeichert: {ausgabe_pfad}")
    return ausgabe_pfad


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Iter-5 XLSX-Export aus KantonsCache",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Beispiele:
    python excel_export.py "Oberhofen am Thunersee"
    python excel_export.py "Thun" --kanton be
    python excel_export.py "Bern" --ausgabe docs/beispiele/
""",
    )
    parser.add_argument("gemeinde", help="Gemeinde-Name (genau wie im Cache)")
    parser.add_argument("--kanton", default=DEFAULT_KANTON, help=f"Kanton-Kuerzel (default: {DEFAULT_KANTON})")
    parser.add_argument(
        "--ausgabe",
        type=Path,
        default=DEFAULT_AUSGABE_DIR,
        help=f"Ausgabe-Verzeichnis (default: {DEFAULT_AUSGABE_DIR})",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=CACHE_DIR,
        help=f"Cache-Verzeichnis (default: {CACHE_DIR})",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    try:
        pfad = exportiere_gemeinde(
            gemeinde=args.gemeinde,
            kanton=args.kanton,
            ausgabe_dir=args.ausgabe,
            cache_dir=args.cache_dir,
        )
        print(f"\nXLSX erstellt: {pfad}")
        print(f"  Absolut: {pfad.resolve()}")
        return 0
    except FileNotFoundError as e:
        print(f"FEHLER: {e}", file=sys.stderr)
        print(f"Tip: erst gemeinde_analyse.py laufen lassen um den Cache zu fuellen.", file=sys.stderr)
        return 1
    except Exception as e:
        logger.exception(f"Fehler: {e}")
        return 2


if __name__ == "__main__":
    sys.exit(main())
